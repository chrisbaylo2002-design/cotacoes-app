"""Gera um Excel (.xlsx) com abas: Hospedagem, Aéreos e Simulador."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cabecalho_folha(ws, titulo, largura_merge="E"):
    ws.merge_cells(f"A1:{largura_merge}1")
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="888888")


def _escrever_tabela(ws, linha_inicio, colunas, dados_linhas, larguras, col_moeda=None):
    for i, col in enumerate(colunas, start=1):
        cel = ws.cell(row=linha_inicio, column=i, value=col)
        cel.fill = HEADER_FILL
        cel.font = HEADER_FONT
        cel.alignment = Alignment(horizontal="center")
        cel.border = BORDER

    row = linha_inicio + 1
    for valores in dados_linhas:
        for col_idx, val in enumerate(valores, start=1):
            cel = ws.cell(row=row, column=col_idx, value=val)
            cel.border = BORDER
            if col_moeda and col_idx in col_moeda:
                cel.number_format = '"R$" #,##0.00'
                cel.alignment = Alignment(horizontal="right")
        row += 1

    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    return row


def gerar_excel(
    nome_fornecedor: str,
    observacoes_gerais: str,
    pivot_rows: list,
    aereos_rows: list,
    simulador_rows: list,
    qtd_apartamentos: float,
    taxa_percentual: float,
) -> io.BytesIO:
    wb = Workbook()

    # --- aba Hospedagem ---
    ws1 = wb.active
    ws1.title = "Hospedagem"
    _cabecalho_folha(ws1, f"Cotação — {nome_fornecedor}", "G")
    linha = 4
    if observacoes_gerais:
        ws1.merge_cells(f"A{linha}:G{linha}")
        ws1[f"A{linha}"] = f"Obs.: {observacoes_gerais}"
        ws1[f"A{linha}"].font = Font(italic=True, size=9)
        linha += 2

    dados_hosp = [
        [
            r["periodo"],
            r["data"],
            r["categoria"],
            r["single"] if r["single"] is not None else "",
            r["duplo"] if r["duplo"] is not None else "",
            r["triplo"] if r["triplo"] is not None else "",
            r.get("observacao", ""),
        ]
        for r in pivot_rows
    ]
    _escrever_tabela(
        ws1,
        linha,
        ["Período", "Data", "Categoria", "Single (R$)", "Duplo (R$)", "Triplo (R$)", "Observação"],
        dados_hosp,
        [30, 22, 30, 14, 14, 14, 30],
        col_moeda={4, 5, 6},
    )

    # --- aba Aéreos ---
    ws2 = wb.create_sheet("Aéreos")
    _cabecalho_folha(ws2, "Aéreos — Ida e Volta", "D")
    dados_aereos = [[r["periodo"], r["nome"], r["cia"], r["valor"]] for r in aereos_rows]
    _escrever_tabela(
        ws2,
        4,
        ["Período", "Passageiro", "Companhia", "Valor (R$)"],
        dados_aereos,
        [35, 22, 22, 16],
        col_moeda={4},
    )

    # --- aba Simulador ---
    ws3 = wb.create_sheet("Simulador de Grupo")
    _cabecalho_folha(ws3, "Simulador de Cotação de Grupo", "G")
    ws3["A3"] = f"{int(qtd_apartamentos)} apartamento(s) — {taxa_percentual:.0f}% de taxas"
    ws3["A3"].font = Font(italic=True, size=9)
    dados_sim = [
        [
            s["periodo"],
            s["categoria"],
            s["ocupacao"].capitalize(),
            s["diaria_media"],
            s["subtotal"],
            s["taxa_valor"],
            s["total"],
        ]
        for s in simulador_rows
    ]
    _escrever_tabela(
        ws3,
        5,
        ["Período", "Categoria", "Ocupação", "Diária média", "Subtotal diárias", "Taxas", "Total do grupo"],
        dados_sim,
        [30, 30, 12, 16, 18, 16, 18],
        col_moeda={4, 5, 6, 7},
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
